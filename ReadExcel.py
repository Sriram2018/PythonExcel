'''
Created on 21-Jan-2018

@author: Sriram
'''
import openpyxl
from openpyxl import load_workbook

# Load in the workbook
wb = load_workbook('E:/Expensehhhs.xlsx')

# Get sheet names
print(wb.sheetnames)

sheet1 = wb["Sheet1"]

row_count = sheet1.max_row
col_count = sheet1.max_column

print(row_count)
print(col_count)

for i in range(1,col_count+1):    
    sheet1['D'+ str(i)]  =  '=IF(B'+str(i)+'<>C'+str(i)+',"0","1")'



wb.save('E:/Expensehhhs.xlsx')

for row in range(1, sheet1.max_row + 1):
    print(sheet1['A'+str(row)].value)
    print(sheet1['B'+str(row)].value)
    print(sheet1['C'+str(row)].value)

    
  


     



